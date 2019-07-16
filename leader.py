#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jul  2 15:45:59 2019

@author: sugiharayuta
"""


import numpy as np
import matplotlib.pyplot as plt
import emcee
import openpyxl
# ブックを取得
book = openpyxl.load_workbook('half earth imp.xlsx')
# シートを取得
sheet = book['Analysis']
# セルを取得
y0=[]
for i in range(3,1030):
    y0.append(sheet.cell(row=i, column=2).value)

y1 = []
for i in y0 :
    j = 90 - i
    y1.append(j)

y2=[]
for i in range(3,1030):
    y2.append(sheet.cell(row=i, column=5).value)

y3 = []
for i in y2 :
    j = i - 90
    y3.append(j)

ydeg =[]
for ind,i in enumerate (y1):
    j = y3 [ind]
    k = (i + j)/2
    ydeg.append(k)

t=[]
for i in range(3,1030):
    t.append(sheet.cell(row=i, column=7).value)

plt.plot(t, ydeg)
plt.show()

yrad = np.radians(ydeg)

yerr = 0.0035

plt.errorbar(t,yrad,yerr, fmt="o")

P = []
t0 = []
e = []
t0e = []
a = []

def f(t,P,t0,e,t0e,a):
    return np.arctan(np.tan(a) * np.sin(2 * np.pi * (t - t0) / P + 2 * e *
                            np.sin(2 * np.pi * (t - t0e) / P) -
                            2 * e * np.sin(2 * np.pi * (t0 - t0e) / P)))

theta = []
def lnlike(theta, t, yrad, yerr):
    # log(likelihood)
    # likelihood = exp(-chi^2/2)
    # log(likelihood) = -chi^2/2
    P,t0,e,t0e,a = theta

    chi2 = np.sum((yrad - f(t,P,t0,e,t0e,a))**2/yerr**2)
    # print "chi2:", chi2
    return -0.5*chi2

ndim, nwalkers = 5, 500

P_init = 365
t0_init = 67
e_init = 0.017
t0e_init = 166
a_init = 0.4
theta_init = np.array([P_init,t0_init,e_init,t0e_init,a_init])

pos = [theta_init + 5e-1*np.random.randn(ndim) for i in range(nwalkers)]

sampler = emcee.EnsembleSampler(nwalkers, ndim, lnlike, threads=10, args=(t, yrad, yerr))
mc = 100
sampler.run_mcmc(pos, mc)
samples = sampler.chain[:, :, :].reshape((-1, ndim))

steps = nwalkers * mc



plt.figure()
plt.plot(samples[:,0]) # a
plt.xlabel("steps")
plt.xlim(0, steps)
plt.ylabel("a")

plt.figure()
plt.plot(samples[:,1]) # b
plt.xlabel("steps")
plt.xlim(0, steps)
plt.ylabel("b")

plt.figure()
plt.plot(samples[:,2]) # b
plt.xlabel("steps")
plt.xlim(0, steps)
plt.ylabel("c")

plt.figure()
plt.plot(samples[:,3]) # b
plt.xlabel("steps")
plt.xlim(0, steps)
plt.ylabel("d")

plt.figure()
plt.plot(samples[:,4]) # b
plt.xlabel("steps")
plt.xlim(0, steps)
plt.ylabel("e")

a_mcmc, b_mcmc, c_mcmc, d_mcmc, e_mcmc= map(lambda v: (v[1], v[2]-v[1], v[1]-v[0]),zip(*np.percentile(samples, [16, 50, 84],axis=0)))
